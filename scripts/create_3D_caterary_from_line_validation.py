import arcpy
from openpyxl import load_workbook

CONDUCTORTABLENAME = "ConductorLUTable.xlsx"
CONDUCTORINSHEET = "ConductorLUTable"
TOWERTABLENAME = "TowerLUTable.xlsx"
TOWERINSHEET = "TowerLUTable"

aprx = arcpy.mp.ArcGISProject("CURRENT")
home_directory = aprx.homeFolder
table_directory = home_directory + "\\tables"
inTowerTable = os.path.join(table_directory, TOWERTABLENAME)
inConductorTable = os.path.join(table_directory, CONDUCTORTABLENAME)


class ToolValidator(object):
    """Class for validating a tool's parameter values and controlling
    the behavior of the tool's dialog."""

    def __init__(self):
        """Setup arcpy and the list of tool parameters."""
        self.params = arcpy.GetParameterInfo()

    def initializeParameters(self):
        """Refine the properties of a tool's parameters. This method is
        called when the tool is opened."""

        # load voltage table
        workbook = load_workbook(inTowerTable)
        worksheet = workbook[TOWERINSHEET]
        voltageList = []

        first_column = worksheet['A']

        # add A column to UI
        for x in range(len(first_column)):
            if first_column[x].value != "Voltage":
                voltageList.append((first_column[x].value))

        self.params[2].filter.list = voltageList
        self.params[3].value = True

        # load conductor table
        workbook = load_workbook(inConductorTable)
        worksheet = workbook[CONDUCTORINSHEET]
        conductorList = []

        first_column = worksheet['A']

        # add A column to UI
        for x in range(len(first_column)):
            if first_column[x].value != "ConductorName":
                conductorList.append((first_column[x].value))

        self.params[4].filter.list = conductorList
        #        self.params[4].value = conductorList[4]
        self.params[5].value = 21500
        self.params[6].value = 15
        self.params[7].value = self.params[5].value * (self.params[6].value / 100)

        self.params[18].value = True
        self.params[18].enabled = False
        self.params[19].value = "Transmission"
        self.params[19].enabled = False
        self.params[21].value = "None"

    def updateParameters(self):
        """Modify the values and properties of parameters before internal
        validation is performed. This method is called whenever a parameter
        has been changed."""

        self.params[18].enabled = False

        if self.params[0].value:
            type = arcpy.Describe(self.params[0].value).shapeType
            if type == "Point":
                self.params[1].enabled = True
            else:
                self.params[1].enabled = False

        circuitList = [1, 2]
        shieldwireList = [0, 1, 2]
        alignmentList = ["Horizontal", "Vertical", "Offset"]
        structureTypeList = ["Lattice", "Pole"]
        t_towerList = ["66kV", "69kV", "110kV", "115kV", "138kV", "161kV", "220kV",
                       "230kV", "345kV", "380kV", "400kV", "500kV"]
        endList = ["None", "Start", "End", "Both"]

        # load conductor table again to find Rate Breaking strength
        workbook = load_workbook(inConductorTable)
        worksheet = workbook[CONDUCTORINSHEET]
        row_number = -99
        rbs_list = []
        name_list = []

        i = 1
        # read conductor name list
        for row_cells in worksheet.iter_rows(min_col=1, max_col=1):
            for cell in row_cells:
                name_list.append(str(cell.value))
                if str(self.params[4].value) == str(cell.value):
                    row_number = i
                i += 1

        j = 1
        # read RBS list
        for row_cells in worksheet.iter_rows(min_col=3, max_col=3):
            for cell in row_cells:
                if str(j) == str(row_number):
                    rbs_list.append(str(cell.value))
                j += 1

        if len(rbs_list) > 0:
            self.params[5].value = rbs_list[0]
            self.params[7].value = self.params[5].value * (self.params[6].value / 100)

        if self.params[3].value == True:
            self.params[4].enabled = False
            self.params[5].enabled = False
            self.params[6].enabled = False
            self.params[7].enabled = False
        else:
            self.params[4].enabled = True
            self.params[5].enabled = True
            self.params[6].enabled = True
            self.params[7].enabled = True

        self.params[13].filter.list = structureTypeList
        self.params[14].filter.list = circuitList
        self.params[15].filter.list = alignmentList
        self.params[17].filter.list = shieldwireList

        if self.params[18].value == True:
            self.params[13].value = structureTypeList[0]
            self.params[14].value = circuitList[1]
            self.params[15].value = alignmentList[1]
            self.params[16].value = "Single"
            self.params[17].value = shieldwireList[1]
            self.params[18].value = False
            self.params[21].value = endList[0]

        self.params[19].value = "Transmission"
        self.params[19].enabled = False
        self.params[20].enabled = False
        self.params[20].value = "Steel"

        self.params[21].filter.list = endList

    def updateMessages(self):
        """Modify the messages created by internal validation for each tool
        parameter. This method is called after internal validation."""

        if self.params[0].value:
            num_Features = int(arcpy.GetCount_management(self.params[0].value).getOutput(0))

            if num_Features == 0:
                self.params[0].setErrorMessage('No lines found! Please check your input feature class.')
            else:
                self.params[0].clearMessage()

        if self.params[3].value == False:
            if self.params[4].value == None:
                self.params[4].setErrorMessage('Please select a conductor!')
            else:
                self.params[4].clearMessage()